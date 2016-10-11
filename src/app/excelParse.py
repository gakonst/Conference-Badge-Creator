from openpyxl import *
from openpyxl.cell.cell import *
import codecs

import sys
sys.stdout = codecs.getwriter('utf-8')(sys.stdout)
def excel_to_dict(file_url):
    #Initialization and Excel file opening
    #file_url = 'C:/Users/GAK/Desktop/test.xlsx'
    print ("[+] Opening Excel worksheet..."),
    wb = load_workbook(filename=file_url)
    ws = wb.active
    print ("Done")
    columns = ws.max_column + 1
    rows = ws.max_row + 1
    names = []
    surnames = []
    occupations = []
    full = []
    #Loop through whole excel file
    print ("[+] Reading worksheet, appending Names/Surnames/Occupations")
    for column in range(1, columns):
        column_letter = get_column_letter(column) #Translate column index to letter
        if column_letter == 'A': #A corresponds to Name
            for row in range(2, rows):
                index = column_letter + str(row) #index = 'AX' where X is the row index
                name = ws[index].value
               # name = name.decode('ISO-8859-1')
                names.append(name)
        if column_letter == 'B': #B corresponds to Surname
            for row in range(2, rows):
                index = column_letter + str(row) #index = 'BX' where X is the row index
                surname = ws[index].value
               # surname = surname.decode('ISO-8859-1')
                surnames.append(surname)
        if column_letter == 'C': #C corresponds to Occupation
            for row in range(2, rows):
                index = column_letter + str(row) #index = 'CX' where X is the row index
                occupation = ws[index].value
                #occupation = occupation.decode('ISO-8859-1')
                occupations.append(occupation)
    #Using dictionary for better clarity, list of lists is fine. Note:
    print ("[+] Writing everything to a list of dictionaries")
    for i in range(0, len(names)): #len(names) is actually number of rows, it's done to iterate through the whole list
        dictionary = {
            "Name": names[i].encode('utf-8'),
            "Surname": surnames[i].encode('utf-8'),
            "Occupation": occupations[i].encode('utf-8'),
        }
       # print '[!] DEBUG'+str(names[i].decode("utf-8"))
        full.append(dictionary)
    print ("[+] Opening file..."),
    attendees = open('Attendees.txt','w')
    print ("Done\n[+] Writing dictionary to file.")
    attendees.write(str(full))
    attendees.close()