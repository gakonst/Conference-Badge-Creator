from openpyxl import *
from openpyxl.cell.cell import *

import sys

DATA_FILE = 'attendees.log'

def append_from_column(lst, ws, index, length):
    [ lst.append( ws[ index+str(i) ].value ) for i in range(length) ]

def parse_data(config):
    with open(config) as f:
        import json
        data = json.loads(f.read())
    return data

def create_dictionary(names, surnames, occupations, universities):
    print ("[+] Writing everything to a list of dictionaries")
    full = []
    for i in range(numOfRows):  
        dictionary = {
            "Name": names[i].encode('utf-8'),
            "Surname": surnames[i].encode('utf-8'),
            "Occupation": occupations[i].encode('utf-8'),
            "University": universities[i].encode('utf-8')
            }
        # print '[!] DEBUG'+str(names[i].decode("utf-8"))
        full.append(dictionary)
    return full

def parse_excelfile(ws, data_format, numOfRows)
    names = []
    surnames = []
    occupations = []
    universities = []

    # Loop through whole excel file
    print ("[+] Reading worksheet, appending Names/Surnames/Occupations")
    for column in range(numOfColumns):
        column_letter = get_column_letter(column)  # openpyxl function
        if data_format[column_letter].lower() == 'name':
            append_from_column(names, ws, column_letter, numOfRows)
        elif data_format[column_letter].lower() == 'surname':  
            append_from_column(surnames, ws, column_letter, numOfRows)
        elif data_format[column_letter].lower() == 'occupation':  
            append_from_column(occupations, ws, column_letter, numOfRows)
        elif data_format[column_letter].lower() == 'university':
            append_from_column(universities, ws, column_letter, numOfRows)
        else:
            print "ERROR!"
            sys.exit(-1)

    return names, surnames, occupations, universities

def excel_to_dict(file_url, config_file):
    ws = load_workbook(filename=file_url).active
    numOfRows = ws.max_row + 1
    numOfColumns = ws.max_column + 1

    data_format = parse_data(config_file)
    names, surnames, occupations, universities = parse_excelfile(ws, data_format, numOfRows)
    dic = createDictionary(names, surnamesm occupations, universities, numOfRows):

    print ("[+] Writing dictionary to file.")
    with open(DATA_FILE, 'w') as f:
        f.write(str(dic))
    print("[!] Done!")

